from __future__ import annotations

from datetime import date, timedelta
import hashlib
import importlib.util
import json
from pathlib import Path
from types import SimpleNamespace

import duckdb
from openpyxl import load_workbook, Workbook
import pytest


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts/run_p4_real_data_validation.py"
SPEC = importlib.util.spec_from_file_location("p4r", SCRIPT)
assert SPEC and SPEC.loader
p4r = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(p4r)


def _sha(value: str) -> str:
    return hashlib.sha256(value.encode()).hexdigest()


def _weekdays() -> list[date]:
    values: list[date] = []
    current = p4r.START
    while current <= p4r.END:
        if current.weekday() < 5:
            values.append(current)
        current += timedelta(days=1)
    return values


def _event_dates(weekdays: list[date]) -> list[date]:
    result = []
    for year in range(2016, 2026):
        for month in (3, 6, 9, 12):
            result.append(next(day for day in weekdays if day.year == year and day.month == month))
    return result


def _dates() -> tuple[list[date], list[date]]:
    weekdays = _weekdays()
    events = _event_dates(weekdays)
    required = {p4r.START, date(2016, 1, 5), p4r.END, *events}
    selected = set(required)
    for day in weekdays:
        if len(selected) == 2514:
            break
        selected.add(day)
    return sorted(selected), events


def _write_distributions(path: Path, events: list[date]) -> list[dict[str, str]]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "TICKER", "CUSIP", "EX-DATE", "RECORD DATE", "PAYABLE DATE",
        "DIVIDEND ($)", "SHORT TERM CAPITAL GAIN ($)",
        "LONG TERM CAPITAL GAIN ($)", "FREQUENCY",
    ])
    rows = []
    for day in events:
        sheet.append([
            "SPY", "78462F103", day.strftime("%m/%d/%Y"),
            day.strftime("%m/%d/%Y"), day.strftime("%m/%d/%Y"),
            1, 0, 0, "Quarterly",
        ])
        rows.append({
            "ex_date": day.isoformat(),
            "record_date": day.isoformat(),
            "pay_date": day.isoformat(),
            "amount": "1",
        })
    workbook.save(path)
    return rows


def _write_nav(path: Path, dates: list[date]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None, None])
    sheet.append([None, "SPY"])
    sheet.append(["Date", "NAV", "Shares Outstanding", "Total Net Assets"])
    for index, day in enumerate(dates):
        sheet.append([day.strftime("%d-%b-%Y"), 100 + index / 100, 10, 1000])
    workbook.save(path)


def _write_database(path: Path, dates: list[date], events: list[date]) -> tuple[str, str]:
    sina_document, tiingo_document = _sha("sina-document"), _sha("tiingo-document")
    event_set = set(events)
    connection = duckdb.connect(str(path))
    connection.execute("CREATE SCHEMA us_equity_research")
    connection.execute("""
        CREATE TABLE us_equity_research.sina_daily_bars (
            snapshot_id VARCHAR, symbol VARCHAR, trade_date DATE, open DOUBLE,
            high DOUBLE, low DOUBLE, close DOUBLE, volume BIGINT,
            source_document_sha256 VARCHAR, available_at TIMESTAMPTZ,
            row_sha256 VARCHAR, quality_status VARCHAR
        )
    """)
    connection.execute("""
        CREATE TABLE us_equity_research.us_daily_total_return_research (
            snapshot_id VARCHAR, symbol VARCHAR, trade_date DATE, open DOUBLE,
            high DOUBLE, low DOUBLE, close DOUBLE, volume BIGINT,
            adj_open DOUBLE, adj_high DOUBLE, adj_low DOUBLE, adj_close DOUBLE,
            div_cash DOUBLE, split_factor DOUBLE, gross_return_factor DOUBLE,
            adjusted_reference_factor DOUBLE, source_document_sha256 VARCHAR,
            available_at TIMESTAMPTZ, availability_basis VARCHAR,
            row_sha256 VARCHAR, quality_status VARCHAR
        )
    """)
    previous = None
    for index, day in enumerate(dates):
        close = 100 + index / 100
        open_price, high, low = close - 0.01, close + 0.02, close - 0.02
        dividend = 1.0 if day in event_set else 0.0
        factor = None if previous is None else (close + dividend) / previous
        adjusted = close + sum(1 for event in events if event <= day)
        sina_hash, tiingo_hash = _sha(f"sina|{day}"), _sha(f"tiingo|{day}")
        connection.execute(
            "INSERT INTO us_equity_research.sina_daily_bars VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            ("sina-test", "SPY", day, open_price, high, low, close, 1000,
             sina_document, None, sina_hash, "PASS"),
        )
        connection.execute(
            "INSERT INTO us_equity_research.us_daily_total_return_research "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            ("tiingo-test", "SPY", day, open_price, high, low, close, 1000,
             adjusted, adjusted, adjusted, adjusted, dividend, 1.0, factor, factor,
             tiingo_document, None, "UNKNOWN", tiingo_hash,
             "PASS_RETROSPECTIVE_ONLY_UNQUALIFIED_PIT"),
        )
        previous = close
    connection.close()
    return sina_document, tiingo_document


def _fixture(tmp_path: Path) -> tuple[Path, Path, Path]:
    data_root = tmp_path / "data"
    snapshot = data_root / "raw/official/spy/ssga/ssga_spy_20260715T052647Z"
    snapshot.mkdir(parents=True)
    dates, events = _dates()
    distribution_path = snapshot / "spdr-etf-historical-distributions.xlsx"
    nav_path = snapshot / "navhist-us-en-spy.xlsx"
    manifest_path = snapshot / "retrieval_manifest.json"
    event_rows = _write_distributions(distribution_path, events)
    _write_nav(nav_path, dates)
    manifest_path.write_text("{}\n", encoding="utf-8")
    database = data_root / "quant_research.duckdb"
    sina_document, tiingo_document = _write_database(database, dates, events)

    definition = json.loads(p4r.DEFINITION.read_text())
    inputs = definition["inputs"]
    inputs["central_database_sha256"] = p4r.sha256_file(database)
    inputs["p3_evidence"]["manifest_sha256"] = p4r.sha256_path(manifest_path)
    inputs["state_street_distributions"]["content_sha256"] = p4r.sha256_path(
        distribution_path
    )
    inputs["state_street_distributions"]["event_identity_sha256"] = p4r.sha256_bytes(
        p4r._compact_json(event_rows)
    )
    inputs["state_street_nav"]["content_sha256"] = p4r.sha256_path(nav_path)
    inputs["sina"].update({
        "snapshot_id": "sina-test",
        "source_document_sha256": sina_document,
        "ordered_row_hash_sha256": p4r.sha256_bytes(
            "".join(_sha(f"sina|{day}") for day in dates).encode()
        ),
    })
    inputs["tiingo"].update({
        "snapshot_id": "tiingo-test",
        "source_document_sha256": tiingo_document,
        "ordered_row_hash_sha256": p4r.sha256_bytes(
            "".join(_sha(f"tiingo|{day}") for day in dates).encode()
        ),
    })
    definition_path = tmp_path / "definition.json"
    definition_path.write_bytes(p4r.canonical_json(definition))
    return definition_path, database, data_root


def test_default_mode_reads_no_database_or_external_file(monkeypatch, capsys):
    monkeypatch.setattr(p4r, "_evaluate", lambda *_: pytest.fail("data was read"))
    assert p4r.main([]) == 0
    value = json.loads(capsys.readouterr().out)
    assert value == {
        "database_read": False,
        "files_written": False,
        "mode": "VALIDATE_ONLY",
        "network_executed": False,
        "research_id": p4r.RESEARCH_ID,
    }


def test_full_synthetic_replay_passes_all_gates(tmp_path):
    definition, database, data_root = _fixture(tmp_path)
    runtime = {"commit": "a" * 40, "tree": "b" * 40, "clean_committed": True}
    report = p4r.build_report(definition, database, data_root, runtime=runtime)
    assert report["system_gate_counts"] == {"passed": 14, "total": 14}
    assert report["system_validation_status"] == "P4_R_RETROSPECTIVE_SYSTEM_VALIDATION_PASS"
    assert report["strategy_evidence_eligible"] is False
    assert report["strict_pit_eligible"] is False
    assert report["portfolio"]["shares"] > 0


def test_database_hash_change_fails_closed(tmp_path):
    definition, database, data_root = _fixture(tmp_path)
    with database.open("ab") as handle:
        handle.write(b"changed")
    with pytest.raises(p4r.ValidationError, match="database SHA-256 changed"):
        p4r.build_report(
            definition, database, data_root,
            runtime={"commit": "a" * 40, "tree": "b" * 40, "clean_committed": True},
        )


def test_forged_p3_result_hash_fails_closed(tmp_path):
    definition_path, database, data_root = _fixture(tmp_path)
    definition = json.loads(definition_path.read_text())
    definition["inputs"]["p3_evidence"]["result_sha256"] = "0" * 64
    definition_path.write_bytes(p4r.canonical_json(definition))
    with pytest.raises(p4r.ValidationError, match="external input hash changed"):
        p4r.build_report(
            definition_path, database, data_root,
            runtime={"commit": "a" * 40, "tree": "b" * 40, "clean_committed": True},
        )


def test_missing_source_row_fails_closed(tmp_path):
    definition_path, database, data_root = _fixture(tmp_path)
    connection = duckdb.connect(str(database))
    connection.execute(
        "DELETE FROM us_equity_research.sina_daily_bars WHERE trade_date = DATE '2017-01-03'"
    )
    connection.close()
    definition = json.loads(definition_path.read_text())
    definition["inputs"]["central_database_sha256"] = p4r.sha256_file(database)
    definition_path.write_bytes(p4r.canonical_json(definition))
    with pytest.raises(p4r.ValidationError, match="row counts changed"):
        p4r.build_report(
            definition_path, database, data_root,
            runtime={"commit": "a" * 40, "tree": "b" * 40, "clean_committed": True},
        )


def test_definition_boundary_change_is_rejected():
    definition = json.loads(p4r.DEFINITION.read_text())
    definition["boundaries"]["strict_pit_eligible"] = True
    with pytest.raises(p4r.ValidationError, match="boundaries changed"):
        p4r._contract(definition)


def test_definition_classification_count_and_tolerance_changes_are_rejected():
    mutations = (
        (lambda value: value.__setitem__("evidence_class", "STRICT_PIT_STRATEGY_EVIDENCE"),
         "classification changed"),
        (lambda value: value["inputs"]["sina"].__setitem__("classification", "PRIMARY"),
         "Sina contract changed"),
        (lambda value: value["inputs"]["sina"].__setitem__("row_count", 1),
         "Sina contract changed"),
        (lambda value: value["tolerances"].__setitem__("raw_ohlc_absolute_usd", "999999"),
         "tolerances changed"),
    )
    for mutate, message in mutations:
        definition = json.loads(p4r.DEFINITION.read_text())
        mutate(definition)
        with pytest.raises(p4r.ValidationError, match=message):
            p4r._contract(definition)


def test_execute_rejects_alternate_definition_path(tmp_path):
    alternate = tmp_path / "definition.json"
    alternate.write_bytes(p4r.DEFINITION.read_bytes())
    with pytest.raises(p4r.ValidationError, match="canonical definition"):
        p4r.main(["--execute", "--definition", str(alternate)])


def test_duplicate_json_key_is_rejected(tmp_path):
    path = tmp_path / "bad.json"
    path.write_text('{"a": 1, "a": 2}')
    with pytest.raises(p4r.ValidationError, match="duplicate JSON key"):
        p4r.strict_load(path)


def test_distribution_amount_outside_tolerance_fails_gate(tmp_path):
    definition_path, database, data_root = _fixture(tmp_path)
    connection = duckdb.connect(str(database))
    connection.execute(
        "UPDATE us_equity_research.us_daily_total_return_research "
        "SET div_cash = div_cash + 0.01 WHERE div_cash > 0"
    )
    connection.close()
    definition = json.loads(definition_path.read_text())
    definition["inputs"]["central_database_sha256"] = p4r.sha256_file(database)
    definition_path.write_bytes(p4r.canonical_json(definition))
    report = p4r.build_report(
        definition_path, database, data_root,
        runtime={"commit": "a" * 40, "tree": "b" * 40, "clean_committed": True},
    )
    assert report["gate_results"]["distribution_cross_source_within_tolerance"] is False
    assert report["system_validation_status"] == "P4_R_RETROSPECTIVE_SYSTEM_VALIDATION_FAIL"


def test_future_pay_date_is_an_exact_pending_receivable(tmp_path):
    definition_path, database, data_root = _fixture(tmp_path)
    definition = json.loads(definition_path.read_text())
    relative = definition["inputs"]["state_street_distributions"]["relative_path"]
    workbook_path = data_root / relative
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    sheet.cell(sheet.max_row, 5, "01/30/2026")
    workbook.save(workbook_path)
    workbook.close()
    rows = p4r._distributions(workbook_path.read_bytes())
    distribution = definition["inputs"]["state_street_distributions"]
    distribution["content_sha256"] = p4r.sha256_path(workbook_path)
    distribution["event_identity_sha256"] = p4r.sha256_bytes(p4r._compact_json(rows))
    definition_path.write_bytes(p4r.canonical_json(definition))
    report = p4r.build_report(
        definition_path, database, data_root,
        runtime={"commit": "a" * 40, "tree": "b" * 40, "clean_committed": True},
    )
    assert report["gate_results"]["distribution_cash_and_nav_accounting_exact"] is True
    assert float(report["portfolio"]["distribution_cash_pending_receivable"]) > 0
    assert report["system_validation_status"] == "P4_R_RETROSPECTIVE_SYSTEM_VALIDATION_PASS"


def test_publish_is_atomic_and_refuses_overwrite(tmp_path):
    output = tmp_path / "result.json"
    digest, sidecar = p4r._publish({"ok": True}, output)
    assert hashlib.sha256(output.read_bytes()).hexdigest() == digest
    assert sidecar.read_text() == f"{digest}  result.json\n"
    with pytest.raises(p4r.ValidationError, match="already exists"):
        p4r._publish({"ok": True}, output)


def test_result_is_published_last_as_commit_marker(tmp_path, monkeypatch):
    output = tmp_path / "result.json"
    sidecar = output.with_suffix(".json.sha256")
    original = p4r.os.replace
    destinations = []

    def observe_order(source, destination):
        destination = Path(destination)
        if destination == output:
            assert sidecar.is_file()
        destinations.append(destination)
        return original(source, destination)

    monkeypatch.setattr(p4r.os, "replace", observe_order)
    p4r._publish({"ok": True}, output)
    assert destinations == [sidecar, output]


def test_publish_rolls_back_if_second_replace_fails(tmp_path, monkeypatch):
    output = tmp_path / "result.json"
    original = p4r.os.replace
    calls = 0

    def fail_second(source, destination):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise OSError("simulated sidecar publication failure")
        return original(source, destination)

    monkeypatch.setattr(p4r.os, "replace", fail_second)
    with pytest.raises(OSError, match="sidecar publication"):
        p4r._publish({"ok": True}, output)
    assert not output.exists()
    assert not output.with_suffix(".json.sha256").exists()
    assert list(tmp_path.iterdir()) == []


def test_pinned_input_rejects_symlink(tmp_path):
    target = tmp_path / "target.json"
    target.write_text("{}\n", encoding="utf-8")
    link = tmp_path / "link.json"
    link.symlink_to(target)
    with pytest.raises(p4r.ValidationError, match="cannot open pinned input"):
        p4r._pinned_bytes(tmp_path, link.name, p4r.sha256_path(target))


def test_dirty_runtime_fails_closed(monkeypatch):
    def fake_run(arguments, **_kwargs):
        if "status" in arguments:
            return SimpleNamespace(stdout="?? result.json\n")
        return SimpleNamespace(stdout="a" * 40 + "\n")

    monkeypatch.setattr(p4r.subprocess, "run", fake_run)
    with pytest.raises(p4r.ValidationError, match="clean committed worktree"):
        p4r._runtime_identity(require_clean=True)


def test_runner_has_no_network_client():
    source = SCRIPT.read_text()
    assert "import requests" not in source
    assert "urllib.request" not in source
    assert "socket." not in source
