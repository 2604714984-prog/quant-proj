"""Outcome-free feasibility smoke for XLG-disclosed S&P 500 Top 50 membership.

The script executes one frozen SEC request graph.  It stops that graph on the
first access failure and only then tests the two frozen Invesco URLs.  It writes
raw responses and hashes under WSL staging, never opens DuckDB, and never
computes returns, factors, portfolio weights, or current rankings.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime, time, timedelta
import hashlib
from html import unescape
from io import BytesIO
import json
from pathlib import Path
import re
import time as time_module
from typing import Any
from urllib.error import HTTPError
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen
import xml.etree.ElementTree as ET
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


RESEARCH_ID = "US_SP500_TOP50_XLG_DISCLOSED_UNIVERSE_V1"
CONTRACT_SHA256 = "1ad3e8c9d3fc7983646391fcec4b2669fc7ba2c78573899d4bf72ba44a7b3977"
PRIOR_RESULT_SHA256 = "36d0581465ea37d26d695bfaa4e1ca1efbc190880b44148058cd30ba712d0877"
USER_AGENT = "quant-proj-research/2.0 (+https://github.com/2604714984-prog/quant-proj)"
SEC_ATOM_URL = (
    "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&"
    "CIK=C000197609&type=NPORT-P&count=100&output=atom"
)
INVESCO_PAGE_URL = (
    "https://www.invesco.com/us/financial-products/etfs/holdings?"
    "audienceType=Investor&ticker=XLG"
)
INVESCO_DOWNLOAD_URL = (
    "https://www.invesco.com/us/financial-products/etfs/holdings/main/holdings/0?"
    "action=download&audienceType=Investor&ticker=XLG"
)
SERIES_ID = "S000060793"
CONTRACT_ID = "C000197609"
REGISTRANT_CIK = "0001209466"
EASTERN = ZoneInfo("America/New_York")
TARGETS = {
    "EARLIEST_QUALIFIED": None,
    "MID_2022": date(2022, 6, 30),
    "START_2025": date(2025, 1, 31),
    "LATEST_COMPLETE_PUBLIC": None,
}
EXCLUDED_TITLE_TOKENS = (
    "PREFERRED",
    "WARRANT",
    "RIGHT",
    "OPTION",
    "FUTURE",
    "SWAP",
    "NOTE",
    "BOND",
    "FUND",
)


class SmokeError(RuntimeError):
    """A frozen input gate failed."""


def _utc_now() -> str:
    return datetime.now(ZoneInfo("UTC")).isoformat()


def _sha256(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False)
        + "\n",
        encoding="utf-8",
    )


def _read_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise SmokeError(f"invalid JSON at {path}") from exc
    if not isinstance(value, dict):
        raise SmokeError(f"JSON root is not an object at {path}")
    return value


def _fetch_once(
    *, url: str, destination: Path, accept: str, timeout_seconds: float = 30.0
) -> tuple[dict[str, Any], bytes]:
    metadata_path = destination.with_suffix(destination.suffix + ".metadata.json")
    if metadata_path.exists():
        metadata = _read_json(metadata_path)
        if metadata.get("url") != url:
            raise SmokeError(f"staging URL collision at {metadata_path}")
        if destination.exists():
            raw = destination.read_bytes()
            if metadata.get("byte_count") != len(raw) or metadata.get(
                "content_sha256"
            ) != _sha256(raw):
                raise SmokeError(f"staging hash mismatch at {destination}")
            return {**metadata, "cache_reused": True}, raw
        if metadata.get("transport_error"):
            return {**metadata, "cache_reused": True}, b""
        raise SmokeError(f"staged response body is missing at {destination}")
    if destination.exists():
        raise SmokeError(f"staged response metadata is missing at {destination}")

    request = Request(
        url,
        headers={"Accept": accept, "User-Agent": USER_AGENT},
        method="GET",
    )
    retrieved_at = _utc_now()
    try:
        with urlopen(request, timeout=timeout_seconds) as response:  # noqa: S310
            status: int | None = int(response.status)
            headers = {str(key).lower(): str(value) for key, value in response.headers.items()}
            raw = response.read()
    except HTTPError as exc:
        status = int(exc.code)
        headers = {str(key).lower(): str(value) for key, value in exc.headers.items()}
        raw = exc.read()
    except OSError as exc:
        metadata = {
            "url": url,
            "retrieved_at": retrieved_at,
            "http_status": None,
            "byte_count": 0,
            "content_sha256": None,
            "content_type": None,
            "content_disposition": None,
            "etag": None,
            "last_modified": None,
            "body_path": None,
            "transport_error": f"{type(exc).__name__}: {exc}",
            "request_user_agent": USER_AGENT,
            "retry_count": 0,
        }
        _write_json(metadata_path, metadata)
        return metadata, b""

    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(raw)
    metadata = {
        "url": url,
        "retrieved_at": retrieved_at,
        "http_status": status,
        "byte_count": len(raw),
        "content_sha256": _sha256(raw),
        "content_type": headers.get("content-type"),
        "content_disposition": headers.get("content-disposition"),
        "etag": headers.get("etag"),
        "last_modified": headers.get("last-modified"),
        "body_path": str(destination),
        "transport_error": None,
        "request_user_agent": USER_AGENT,
        "retry_count": 0,
    }
    _write_json(metadata_path, metadata)
    return metadata, raw


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _first_descendant_text(element: ET.Element, *names: str) -> str | None:
    wanted = set(names)
    for child in element.iter():
        if _local_name(child.tag) in wanted and child.text and child.text.strip():
            return child.text.strip()
    return None


def _all_descendant_text(element: ET.Element, name: str) -> list[str]:
    return [
        child.text.strip()
        for child in element.iter()
        if _local_name(child.tag) == name and child.text and child.text.strip()
    ]


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    match = re.search(r"\d{4}-\d{2}-\d{2}", value)
    return date.fromisoformat(match.group(0)) if match else None


def _parse_atom_entries(raw: bytes) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as exc:
        raise SmokeError("SEC Atom route did not return valid XML") from exc
    entries: list[dict[str, Any]] = []
    for entry in root.iter():
        if _local_name(entry.tag) != "entry":
            continue
        fields: dict[str, str] = {}
        links: list[str] = []
        for child in entry.iter():
            name = _local_name(child.tag)
            text = child.text.strip() if child.text and child.text.strip() else None
            if text and name not in fields:
                fields[name] = text
            href = child.attrib.get("href")
            if href:
                links.append(href)
        filing_href = fields.get("filing-href")
        if not filing_href:
            filing_href = next(
                (link for link in links if "Archives/edgar/data" in link and "-index" in link),
                None,
            )
        accession = fields.get("accession-number")
        form = fields.get("filing-type") or fields.get("category")
        period = _parse_iso_date(fields.get("period") or fields.get("report-date"))
        filing_date = _parse_iso_date(fields.get("filing-date") or fields.get("updated"))
        if not accession or not filing_href or not form:
            continue
        entries.append(
            {
                "accession_number": accession,
                "form": form,
                "period": period,
                "filing_date": filing_date,
                "updated": fields.get("updated"),
                "filing_href": filing_href,
                "primary_document": fields.get("primary-document"),
            }
        )
    if not entries:
        raise SmokeError("SEC Atom route contained no parseable filing entries")
    return entries


def _entry_sort_key(entry: dict[str, Any]) -> tuple[date, str, str]:
    return (
        entry.get("filing_date") or date.min,
        str(entry.get("updated") or ""),
        str(entry["accession_number"]),
    )


def _select_four_entries(entries: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    eligible = [
        entry
        for entry in entries
        if entry.get("period") is not None
        and entry["period"] >= date(2019, 9, 1)
        and str(entry["form"]).upper() in {"NPORT-P", "NPORT-P/A"}
    ]
    if not eligible:
        raise SmokeError("SEC Atom entries have no eligible report periods")
    by_period: dict[date, list[dict[str, Any]]] = {}
    for entry in eligible:
        by_period.setdefault(entry["period"], []).append(entry)
    latest_revision_by_period = {
        period: max(period_entries, key=_entry_sort_key)
        for period, period_entries in by_period.items()
    }
    periods = sorted(latest_revision_by_period)

    def latest_not_after(cutoff: date) -> date:
        candidates = [period for period in periods if period <= cutoff]
        if not candidates:
            raise SmokeError(f"no SEC report period on or before {cutoff}")
        return max(candidates)

    selected_periods = {
        "EARLIEST_QUALIFIED": min(periods),
        "MID_2022": latest_not_after(TARGETS["MID_2022"]),
        "START_2025": latest_not_after(TARGETS["START_2025"]),
        "LATEST_COMPLETE_PUBLIC": max(periods),
    }
    if len(set(selected_periods.values())) != 4:
        raise SmokeError("the four frozen selectors did not resolve to four distinct periods")
    return {
        selector: {
            **latest_revision_by_period[period],
            "selector": selector,
            "revision_count_for_period": len(by_period[period]),
        }
        for selector, period in selected_periods.items()
    }


def _require_sec_archive_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or parsed.netloc.lower() not in {
        "www.sec.gov",
        "sec.gov",
    }:
        raise SmokeError(f"non-SEC filing URL rejected: {url}")
    if "/Archives/edgar/data/1209466/" not in parsed.path:
        raise SmokeError(f"unexpected registrant archive path: {url}")
    return parsed._replace(scheme="https", netloc="www.sec.gov").geturl()


def _parse_index_html(raw: bytes, index_url: str) -> tuple[str, str]:
    text = raw.decode("utf-8", errors="replace")
    accepted_match = re.search(
        r"Accepted\s*</div>\s*<div[^>]*>\s*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not accepted_match:
        accepted_match = re.search(
            r"Accepted[^0-9]{0,200}(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})",
            text,
            flags=re.IGNORECASE | re.DOTALL,
        )
    if not accepted_match:
        raise SmokeError("SEC filing index has no exact Accepted timestamp")
    accepted_at = datetime.strptime(
        accepted_match.group(1), "%Y-%m-%d %H:%M:%S"
    ).replace(tzinfo=EASTERN)

    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", text, flags=re.IGNORECASE | re.DOTALL)
    xml_href: str | None = None
    for row in rows:
        if not re.search(r">\s*NPORT-P(?:/A)?\s*<", row, flags=re.IGNORECASE):
            continue
        hrefs = re.findall(r'href=["\']([^"\']+\.xml)["\']', row, flags=re.IGNORECASE)
        if hrefs:
            xml_href = hrefs[0]
            break
    if xml_href is None:
        hrefs = re.findall(r'href=["\']([^"\']+\.xml)["\']', text, flags=re.IGNORECASE)
        candidates = [href for href in hrefs if "xsl" not in href.lower()]
        if len(candidates) == 1:
            xml_href = candidates[0]
    if xml_href is None:
        raise SmokeError("SEC filing index has no unique primary NPORT XML link")
    return accepted_at.isoformat(), _require_sec_archive_url(urljoin(index_url, xml_href))


def _normalized_cusip(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"[^A-Z0-9]", "", value.upper())
    return normalized if len(normalized) == 9 else None


def _normalized_name(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"[^A-Z0-9]+", " ", unescape(value).upper()).strip()
    return normalized or None


def _holding_ticker(element: ET.Element) -> str | None:
    for child in element.iter():
        if _local_name(child.tag).lower() != "ticker":
            continue
        value = child.attrib.get("value") or child.text
        if value and value.strip():
            return value.strip()
    return None


def _parse_nport_xml(raw: bytes) -> dict[str, Any]:
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as exc:
        raise SmokeError("primary NPORT document is not valid XML") from exc
    series_ids = sorted(set(_all_descendant_text(root, "seriesId")))
    contract_ids = sorted(set(_all_descendant_text(root, "classId")))
    if SERIES_ID not in series_ids or CONTRACT_ID not in contract_ids:
        raise SmokeError(
            f"NPORT identity mismatch: series={series_ids} contracts={contract_ids}"
        )
    cik_values = {
        re.sub(r"\D", "", value).zfill(10)
        for value in _all_descendant_text(root, "cik")
        if re.sub(r"\D", "", value)
    }
    if REGISTRANT_CIK not in cik_values:
        raise SmokeError(f"NPORT registrant CIK mismatch: {sorted(cik_values)}")
    report_date_text = _first_descendant_text(
        root, "repPdDate", "reportDate", "periodOfReport"
    )
    report_date = _parse_iso_date(report_date_text)
    if report_date is None:
        raise SmokeError("NPORT document has no report date")

    holdings: list[dict[str, Any]] = []
    for element in root.iter():
        if _local_name(element.tag) != "invstOrSec":
            continue
        name = _first_descendant_text(element, "name")
        title = _first_descendant_text(element, "title")
        cusip = _normalized_cusip(_first_descendant_text(element, "cusip"))
        asset_category = _first_descendant_text(element, "assetCat")
        ticker = _holding_ticker(element)
        normalized_title = _normalized_name(title) or ""
        included = (
            asset_category == "EC"
            and cusip is not None
            and name is not None
            and title is not None
            and ticker is not None
            and not any(token in normalized_title for token in EXCLUDED_TITLE_TOKENS)
        )
        holdings.append(
            {
                "cusip": cusip,
                "issuer_key": cusip[:6] if cusip else None,
                "name_normalized": _normalized_name(name),
                "title_normalized": normalized_title or None,
                "ticker_present": ticker is not None,
                "asset_category": asset_category,
                "included_common_equity": included,
            }
        )
    if not holdings:
        raise SmokeError("NPORT document has no investment-or-security rows")
    included = [holding for holding in holdings if holding["included_common_equity"]]
    company_names: dict[str, set[str]] = {}
    for holding in included:
        issuer_key = str(holding["issuer_key"])
        company_names.setdefault(issuer_key, set()).add(str(holding["name_normalized"]))
    name_conflicts = {
        hashlib.sha256(issuer_key.encode()).hexdigest(): len(names)
        for issuer_key, names in company_names.items()
        if len(names) > 1
    }
    security_key_counts = Counter(
        (holding["cusip"], holding["title_normalized"]) for holding in included
    )
    duplicate_security_keys = sum(count - 1 for count in security_key_counts.values())
    line_counts_by_company = Counter(str(holding["issuer_key"]) for holding in included)
    multi_class_company_hashes = sorted(
        hashlib.sha256(issuer_key.encode()).hexdigest()
        for issuer_key, count in line_counts_by_company.items()
        if count > 1
    )
    asset_category_counts = Counter(
        str(holding["asset_category"] or "MISSING") for holding in holdings
    )
    return {
        "report_date": report_date.isoformat(),
        "series_id_match": True,
        "contract_id_match": True,
        "registrant_cik_match": True,
        "all_investment_rows": len(holdings),
        "eligible_common_equity_lines": len(included),
        "company_count": len(line_counts_by_company),
        "multi_class_company_count": len(multi_class_company_hashes),
        "multi_class_company_key_sha256": multi_class_company_hashes,
        "duplicate_security_keys": duplicate_security_keys,
        "issuer_name_conflict_count": len(name_conflicts),
        "issuer_name_conflict_key_sha256": sorted(name_conflicts),
        "asset_category_counts": dict(sorted(asset_category_counts.items())),
        "excluded_investment_rows": len(holdings) - len(included),
    }


def _observed(day: date) -> date:
    if day.weekday() == 5:
        return day - timedelta(days=1)
    if day.weekday() == 6:
        return day + timedelta(days=1)
    return day


def _nth_weekday(year: int, month: int, weekday: int, ordinal: int) -> date:
    first = date(year, month, 1)
    offset = (weekday - first.weekday()) % 7
    return first + timedelta(days=offset + 7 * (ordinal - 1))


def _last_weekday(year: int, month: int, weekday: int) -> date:
    first_next = date(year + (month == 12), 1 if month == 12 else month + 1, 1)
    candidate = first_next - timedelta(days=1)
    return candidate - timedelta(days=(candidate.weekday() - weekday) % 7)


def _easter_sunday(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    ell = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * ell) // 451
    month = (h + ell - 7 * m + 114) // 31
    day = (h + ell - 7 * m + 114) % 31 + 1
    return date(year, month, day)


def _xnys_holidays(year: int) -> set[date]:
    holidays = {
        _observed(date(year, 1, 1)),
        _nth_weekday(year, 1, 0, 3),
        _nth_weekday(year, 2, 0, 3),
        _easter_sunday(year) - timedelta(days=2),
        _last_weekday(year, 5, 0),
        _observed(date(year, 7, 4)),
        _nth_weekday(year, 9, 0, 1),
        _nth_weekday(year, 11, 3, 4),
        _observed(date(year, 12, 25)),
    }
    if year >= 2022:
        holidays.add(_observed(date(year, 6, 19)))
    return holidays


EARLY_CLOSES = {
    date.fromisoformat(value)
    for value in (
        "2019-07-03",
        "2019-11-29",
        "2019-12-24",
        "2020-11-27",
        "2020-12-24",
        "2021-11-26",
        "2022-11-25",
        "2023-07-03",
        "2023-11-24",
        "2024-07-03",
        "2024-11-29",
        "2024-12-24",
        "2025-07-03",
        "2025-11-28",
        "2025-12-24",
        "2026-07-02",
        "2026-11-27",
        "2026-12-24",
    )
}


def _next_complete_xnys_open(accepted_at: str) -> str:
    accepted = datetime.fromisoformat(accepted_at).astimezone(EASTERN)
    candidate = accepted.date() + timedelta(days=1)
    while (
        candidate.weekday() >= 5
        or candidate in _xnys_holidays(candidate.year)
        or candidate in EARLY_CLOSES
    ):
        candidate += timedelta(days=1)
    return datetime.combine(candidate, time(9, 30), EASTERN).isoformat()


def _sec_snapshot_gate(summary: dict[str, Any], selected_period: date) -> list[str]:
    failures: list[str] = []
    if summary["report_date"] != selected_period.isoformat():
        failures.append("report_date_mismatch")
    if summary["company_count"] != 50:
        failures.append("company_count_not_50_without_deletion_evidence")
    if summary["eligible_common_equity_lines"] not in {50, 51}:
        failures.append("eligible_security_line_count_not_50_or_51")
    if summary["duplicate_security_keys"] != 0:
        failures.append("duplicate_security_keys")
    if summary["issuer_name_conflict_count"] != 0:
        failures.append("issuer_name_conflicts")
    if summary["eligible_common_equity_lines"] - summary["company_count"] != summary[
        "multi_class_company_count"
    ]:
        failures.append("unexplained_multi_class_line_count")
    if summary["multi_class_company_count"]:
        failures.append("trailing_60_session_dollar_volume_not_qualified_in_this_smoke")
    return failures


def _run_sec(staging: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    documents: list[dict[str, Any]] = []
    atom_metadata, atom_raw = _fetch_once(
        url=SEC_ATOM_URL,
        destination=staging / "sec" / "xlg_contract_nport_atom.xml",
        accept="application/atom+xml,application/xml,text/xml;q=0.9,*/*;q=0.1",
    )
    documents.append(atom_metadata)
    status = atom_metadata["http_status"]
    if status != 200:
        return (
            {
                "route": "SEC_EDGAR_NPORT_XML",
                "status": "ACCESS_BLOCKED" if status == 403 else "SOURCE_FAILURE",
                "initial_http_status": status,
                "initial_transport_error": atom_metadata.get("transport_error"),
                "request_count": 1,
                "retry_count": 0,
                "header_identity_changed": False,
                "four_snapshot_pass": False,
                "blocking_failures": ["sec_initial_access_not_200"],
            },
            documents,
        )

    entries = _parse_atom_entries(atom_raw)
    selected = _select_four_entries(entries)
    snapshot_results: list[dict[str, Any]] = []
    request_count = 1
    for selector, entry in selected.items():
        time_module.sleep(1.1)
        index_url = _require_sec_archive_url(str(entry["filing_href"]))
        index_metadata, index_raw = _fetch_once(
            url=index_url,
            destination=staging / "sec" / f"{selector.lower()}_index.html",
            accept="text/html,application/xhtml+xml;q=0.9,*/*;q=0.1",
        )
        documents.append(index_metadata)
        request_count += 1
        if index_metadata["http_status"] != 200:
            return (
                {
                    "route": "SEC_EDGAR_NPORT_XML",
                    "status": "ACCESS_BLOCKED"
                    if index_metadata["http_status"] == 403
                    else "SOURCE_FAILURE",
                    "initial_http_status": 200,
                    "request_count": request_count,
                    "retry_count": 0,
                    "header_identity_changed": False,
                    "four_snapshot_pass": False,
                    "blocking_failures": [f"{selector.lower()}_index_not_200"],
                    "snapshots_completed": len(snapshot_results),
                },
                documents,
            )
        accepted_at, xml_url = _parse_index_html(index_raw, index_url)

        time_module.sleep(1.1)
        xml_metadata, xml_raw = _fetch_once(
            url=xml_url,
            destination=staging / "sec" / f"{selector.lower()}_primary.xml",
            accept="application/xml,text/xml;q=0.9,*/*;q=0.1",
        )
        documents.append(xml_metadata)
        request_count += 1
        if xml_metadata["http_status"] != 200:
            return (
                {
                    "route": "SEC_EDGAR_NPORT_XML",
                    "status": "ACCESS_BLOCKED"
                    if xml_metadata["http_status"] == 403
                    else "SOURCE_FAILURE",
                    "initial_http_status": 200,
                    "request_count": request_count,
                    "retry_count": 0,
                    "header_identity_changed": False,
                    "four_snapshot_pass": False,
                    "blocking_failures": [f"{selector.lower()}_primary_xml_not_200"],
                    "snapshots_completed": len(snapshot_results),
                },
                documents,
            )
        summary = _parse_nport_xml(xml_raw)
        period = entry["period"]
        failures = _sec_snapshot_gate(summary, period)
        snapshot_results.append(
            {
                "selector": selector,
                "accession_number": entry["accession_number"],
                "form": entry["form"],
                "selected_period": period.isoformat(),
                "accepted_at": accepted_at,
                "effective_at": _next_complete_xnys_open(accepted_at),
                "revision_count_for_period": entry["revision_count_for_period"],
                "index_sha256": index_metadata["content_sha256"],
                "primary_xml_sha256": xml_metadata["content_sha256"],
                "summary": summary,
                "gate_failures": failures,
                "pass": not failures,
            }
        )
    blocking = sorted(
        {
            failure
            for snapshot in snapshot_results
            for failure in snapshot["gate_failures"]
        }
    )
    return (
        {
            "route": "SEC_EDGAR_NPORT_XML",
            "status": "PASS" if not blocking else "QUALITY_FAIL",
            "initial_http_status": 200,
            "request_count": request_count,
            "retry_count": 0,
            "header_identity_changed": False,
            "atom_entry_count": len(entries),
            "snapshot_count": len(snapshot_results),
            "snapshots": snapshot_results,
            "four_snapshot_pass": len(snapshot_results) == 4 and not blocking,
            "blocking_failures": blocking,
        },
        documents,
    )


def _inspect_invesco_download(raw: bytes, content_type: str | None) -> dict[str, Any]:
    format_name = "unknown"
    row_count: int | None = None
    column_count: int | None = None
    text_sample = ""
    lowered_type = (content_type or "").lower()
    decoded_prefix = raw[:200_000].decode("utf-8", errors="replace")
    if "html" in lowered_type or "<html" in decoded_prefix.lower():
        format_name = "html"
        text_sample = decoded_prefix
    elif raw.startswith(b"PK\x03\x04"):
        format_name = "xlsx"
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        row_count = sheet.max_row
        column_count = sheet.max_column
        text_sample = "\n".join(
            " | ".join(str(value) for value in row if value is not None)
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True)
        )
        workbook.close()
    elif raw.startswith(b"\xd0\xcf\x11\xe0"):
        format_name = "xls_ole"
    else:
        text_sample = decoded_prefix
        if "csv" in lowered_type or "," in text_sample[:1000]:
            format_name = "csv_or_text"
            lines = [line for line in text_sample.splitlines() if line.strip()]
            row_count = len(lines)
            column_count = max((len(line.split(",")) for line in lines), default=0)
    date_tokens = sorted(set(re.findall(r"\b(?:20\d{2})[-/]\d{1,2}[-/]\d{1,2}\b", text_sample)))
    publication_tokens = re.findall(
        r"(?:published|publication\s+date|posted\s+at)\s*[:\-]?\s*[^\n<]{0,80}",
        text_sample,
        flags=re.IGNORECASE,
    )
    return {
        "format": format_name,
        "row_count": row_count,
        "column_count": column_count,
        "date_token_count": len(date_tokens),
        "publication_timestamp_token_count": len(publication_tokens),
    }


def _run_invesco(staging: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    documents: list[dict[str, Any]] = []
    page_metadata, page_raw = _fetch_once(
        url=INVESCO_PAGE_URL,
        destination=staging / "invesco" / "xlg_holdings_page.html",
        accept="text/html,application/xhtml+xml;q=0.9,*/*;q=0.1",
    )
    documents.append(page_metadata)
    if page_metadata["http_status"] != 200:
        return (
            {
                "route": "INVESCO_OFFICIAL_HOLDINGS",
                "status": "SOURCE_FAILURE",
                "request_count": 1,
                "retry_count": 0,
                "four_snapshot_pass": False,
                "blocking_failures": ["invesco_holdings_page_not_200"],
                "page_http_status": page_metadata["http_status"],
            },
            documents,
        )
    download_metadata, download_raw = _fetch_once(
        url=INVESCO_DOWNLOAD_URL,
        destination=staging / "invesco" / "xlg_holdings_download.bin",
        accept=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/vnd.ms-excel,text/csv;q=0.9,*/*;q=0.1"
        ),
    )
    documents.append(download_metadata)
    if download_metadata["http_status"] != 200:
        return (
            {
                "route": "INVESCO_OFFICIAL_HOLDINGS",
                "status": "SOURCE_FAILURE",
                "request_count": 2,
                "retry_count": 0,
                "four_snapshot_pass": False,
                "blocking_failures": ["invesco_holdings_download_not_200"],
                "page_http_status": 200,
                "download_http_status": download_metadata["http_status"],
            },
            documents,
        )
    page_text = page_raw.decode("utf-8", errors="replace")
    historical_markers = re.findall(
        r"historical\s+(?:portfolio\s+)?holdings|holdings\s+history",
        page_text,
        flags=re.IGNORECASE,
    )
    publication_markers = re.findall(
        r"(?:published|publication\s+date|posted\s+at)\s*[:\-]?\s*[^\n<]{0,80}",
        page_text,
        flags=re.IGNORECASE,
    )
    inspection = _inspect_invesco_download(
        download_raw, str(download_metadata.get("content_type") or "")
    )
    failures: list[str] = []
    if not historical_markers:
        failures.append("no_official_historical_holdings_archive_identity")
    if not publication_markers and inspection["publication_timestamp_token_count"] == 0:
        failures.append("no_exact_publication_timestamp_bound_to_holdings_bytes")
    failures.append("four_frozen_historical_snapshots_not_exposed_by_frozen_urls")
    return (
        {
            "route": "INVESCO_OFFICIAL_HOLDINGS",
            "status": "QUALITY_FAIL",
            "request_count": 2,
            "retry_count": 0,
            "page_http_status": 200,
            "download_http_status": 200,
            "historical_archive_marker_count": len(historical_markers),
            "publication_timestamp_marker_count": len(publication_markers),
            "download_inspection": inspection,
            "snapshot_count": 1 if inspection["row_count"] else 0,
            "four_snapshot_pass": False,
            "blocking_failures": sorted(set(failures)),
        },
        documents,
    )


def run(contract_path: Path, prior_result_path: Path, staging: Path) -> dict[str, Any]:
    contract_raw = contract_path.read_bytes()
    prior_raw = prior_result_path.read_bytes()
    if _sha256(contract_raw) != CONTRACT_SHA256:
        raise SmokeError("new identity contract hash mismatch")
    if _sha256(prior_raw) != PRIOR_RESULT_SHA256:
        raise SmokeError("prior INPUT_BLOCKED result hash mismatch")
    contract = _read_json(contract_path)
    if contract.get("research_id") != RESEARCH_ID:
        raise SmokeError("new identity research_id mismatch")

    sec_result, sec_documents = _run_sec(staging)
    source_documents = list(sec_documents)
    invesco_result: dict[str, Any] | None = None
    fallback_allowed = sec_result["status"] == "ACCESS_BLOCKED"
    if fallback_allowed:
        invesco_result, invesco_documents = _run_invesco(staging)
        source_documents.extend(invesco_documents)

    sec_pass = bool(sec_result.get("four_snapshot_pass"))
    invesco_pass = bool(invesco_result and invesco_result.get("four_snapshot_pass"))
    route_feasible = sec_pass or invesco_pass
    usable_interval: dict[str, str] | None = None
    if sec_pass:
        snapshots = sec_result["snapshots"]
        usable_interval = {
            "start_effective_at": min(snapshot["effective_at"] for snapshot in snapshots),
            "end_report_date": max(
                snapshot["summary"]["report_date"] for snapshot in snapshots
            ),
            "basis": "four qualified SEC NPORT snapshots",
        }

    result = {
        "research_id": RESEARCH_ID,
        "date": "2026-07-22",
        "market": "US_EQUITIES",
        "phase": "outcome_free_input_route_feasibility",
        "status": "DATA_ROUTE_FEASIBLE" if route_feasible else "INPUT_BLOCKED",
        "current": True,
        "strategy_candidate_available": False,
        "contract_path": str(contract_path),
        "contract_sha256": CONTRACT_SHA256,
        "prior_result_path": str(prior_result_path),
        "prior_result_sha256_before": PRIOR_RESULT_SHA256,
        "prior_result_sha256_after": _sha256(prior_result_path.read_bytes()),
        "prior_result_preserved": _sha256(prior_result_path.read_bytes())
        == PRIOR_RESULT_SHA256,
        "universe_identity": RESEARCH_ID,
        "series_id": SERIES_ID,
        "contract_id": CONTRACT_ID,
        "sec_route": sec_result,
        "invesco_fallback_attempted": invesco_result is not None,
        "invesco_route": invesco_result,
        "usable_interval": usable_interval,
        "source_documents": source_documents,
        "source_document_count": len(source_documents),
        "central_database_opened": False,
        "central_database_write": False,
        "return_computation": False,
        "factor_computation": False,
        "portfolio_computation": False,
        "alpha_preregistration": False,
        "validation_opened": False,
        "holdout_opened": False,
        "current_constituents_emitted": False,
        "fund_weights_used_as_portfolio_weights": False,
        "boundary_result": (
            "DATA_ROUTE_FEASIBLE_AWAIT_MATERIALIZATION"
            if route_feasible
            else "INPUT_BLOCKED_NO_CENTRAL_WRITE_NO_OUTCOME_ACCESS"
        ),
    }
    _write_json(staging / "smoke_result.json", result)
    _write_json(
        staging / "source_manifest.json",
        {
            "research_id": RESEARCH_ID,
            "generated_at": _utc_now(),
            "source_terms": [
                "https://www.sec.gov/privacy",
                "https://www.sec.gov/developer",
                "https://www.invesco.com/us/en/terms-of-use.html",
            ],
            "source_documents": source_documents,
            "raw_document_count": len(source_documents),
            "raw_byte_count": sum(int(document["byte_count"]) for document in source_documents),
            "retry_count": 0,
        },
    )
    return result


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--contract",
        type=Path,
        default=Path("research/definitions/us_sp500_top50_xlg_disclosed_universe_v1.json"),
    )
    parser.add_argument(
        "--prior-result",
        type=Path,
        default=Path("research/results/us_stock_top50_input_blocked_20260722.json"),
    )
    parser.add_argument("--staging-dir", type=Path, required=True)
    parser.add_argument("--execute-network-smoke", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    if not args.execute_network_smoke:
        raise SystemExit("--execute-network-smoke is required")
    staging = args.staging_dir.resolve()
    try:
        result = run(args.contract.resolve(), args.prior_result.resolve(), staging)
    except SmokeError as exc:
        result = {
            "research_id": RESEARCH_ID,
            "phase": "outcome_free_input_route_feasibility",
            "status": "INPUT_BLOCKED",
            "current": True,
            "strategy_candidate_available": False,
            "fatal_error": f"{type(exc).__name__}: {exc}",
            "central_database_opened": False,
            "central_database_write": False,
            "return_computation": False,
            "portfolio_computation": False,
            "boundary_result": "INPUT_BLOCKED_NO_CENTRAL_WRITE_NO_OUTCOME_ACCESS",
        }
        _write_json(staging / "smoke_result.json", result)
    print(
        json.dumps(
            {
                "status": result["status"],
                "result_path": str(staging / "smoke_result.json"),
            },
            sort_keys=True,
            separators=(",", ":"),
        )
    )
    return 0 if result["status"] == "DATA_ROUTE_FEASIBLE" else 2


if __name__ == "__main__":
    raise SystemExit(main())
