#!/usr/bin/env python3
"""Qualify one locally preserved snapshot of the fixed SPY official sources."""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
from io import BytesIO
import json
import os
from pathlib import Path
import tempfile
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFINITION = ROOT / "research/definitions/p3_spy_official_source_qualification_v1.json"
MAX_XLSX_UNCOMPRESSED_BYTES = 40_000_000


class QualificationError(ValueError):
    """Raised when official source bytes or their semantics fail closed."""


def _pairs(values: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in values:
        if key in result:
            raise QualificationError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _nonfinite(value: str) -> None:
    raise QualificationError(f"nonfinite JSON constant: {value}")


def strict_load_bytes(raw: bytes) -> dict[str, Any]:
    try:
        value = json.loads(
            raw.decode("utf-8"),
            object_pairs_hook=_pairs,
            parse_constant=_nonfinite,
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise QualificationError("invalid UTF-8 JSON") from exc
    if type(value) is not dict:
        raise QualificationError("top-level JSON value must be an object")
    return value


def strict_load(path: Path) -> dict[str, Any]:
    return strict_load_bytes(path.read_bytes())


def canonical_json(value: Any) -> bytes:
    return (
        json.dumps(value, allow_nan=False, ensure_ascii=True, indent=2, sort_keys=True) + "\n"
    ).encode()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _decimal(value: Any, field: str, *, positive: bool = False) -> Decimal:
    if isinstance(value, bool):
        raise QualificationError(f"{field} must be finite")
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise QualificationError(f"{field} must be finite") from exc
    if not parsed.is_finite() or (positive and parsed <= 0):
        raise QualificationError(f"{field} must be finite and valid")
    return parsed


def _day(value: Any, formats: tuple[str, ...], field: str) -> date:
    if type(value) is datetime:
        return value.date()
    if type(value) is date:
        return value
    for pattern in formats:
        try:
            return datetime.strptime(str(value).strip(), pattern).date()
        except ValueError:
            continue
    raise QualificationError(f"invalid {field}: {value!r}")


def _timestamp(value: Any, field: str) -> datetime:
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError as exc:
        raise QualificationError(f"invalid {field}") from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise QualificationError(f"{field} must be timezone-aware")
    return parsed


def xlsx_rows(raw: bytes) -> list[list[Any]]:
    try:
        with ZipFile(BytesIO(raw)) as archive:
            if sum(member.file_size for member in archive.infolist()) > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise QualificationError("XLSX uncompressed size exceeds the fixed ceiling")
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except (BadZipFile, OSError, ValueError) as exc:
        raise QualificationError("invalid XLSX container") from exc
    try:
        if len(workbook.sheetnames) != 1:
            raise QualificationError("official workbook must contain exactly one worksheet")
        return [list(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _header(rows: list[list[Any]], required: tuple[str, ...]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows):
        columns = {_text(value): index for index, value in enumerate(row) if _text(value)}
        if all(name in columns for name in required):
            return row_index, columns
    raise QualificationError(f"XLSX header is missing: {required}")


def _cell(row: list[Any], columns: dict[str, int], name: str) -> Any:
    index = columns[name]
    return row[index] if index < len(row) else None


def qualify_product(raw: bytes) -> dict[str, Any]:
    rows = xlsx_rows(raw)
    required = ("As of**", "Ticker", "ISIN", "CUSIP", "Inception Date")
    header_index, columns = _header(rows, required)
    matches = [row for row in rows[header_index + 1 :] if _text(_cell(row, columns, "Ticker")) == "SPY"]
    if len(matches) != 1:
        raise QualificationError("product snapshot must contain exactly one SPY row")
    row = matches[0]
    expected = {
        "Ticker": "SPY",
        "ISIN": "US78462F1030",
        "CUSIP": "78462F103",
        "Inception Date": "Jan 22 1993",
        "Distribution Frequency": "Quarterly",
        "Primary Exchange": "NYSE ARCA",
    }
    for field, expected_value in expected.items():
        if field not in columns or _text(_cell(row, columns, field)) != expected_value:
            raise QualificationError(f"SPY product identity mismatch: {field}")
    return {
        "classification": "QUALIFIED_PRIMARY_STATIC_IDENTITY",
        "snapshot_date": _day(
            _cell(row, columns, "As of**"), ("%b %d %Y",), "product date"
        ).isoformat(),
        "ticker": "SPY",
        "cusip": "78462F103",
        "isin": "US78462F1030",
        "primary_exchange": "NYSE ARCA",
        "inception_date": "1993-01-22",
    }


def qualify_distributions(raw: bytes, start: date, end: date) -> dict[str, Any]:
    rows = xlsx_rows(raw)
    required = (
        "TICKER",
        "CUSIP",
        "EX-DATE",
        "RECORD DATE",
        "PAYABLE DATE",
        "DIVIDEND ($)",
        "SHORT TERM CAPITAL GAIN ($)",
        "LONG TERM CAPITAL GAIN ($)",
        "FREQUENCY",
    )
    header_index, columns = _header(rows, required)
    all_dates: list[date] = []
    target_dates: list[date] = []
    for row in rows[header_index + 1 :]:
        if _text(_cell(row, columns, "TICKER")) != "SPY":
            continue
        if _text(_cell(row, columns, "CUSIP")) != "78462F103":
            raise QualificationError("SPY distribution CUSIP mismatch")
        ex_date = _day(_cell(row, columns, "EX-DATE"), ("%m/%d/%Y",), "ex-date")
        record_date = _day(
            _cell(row, columns, "RECORD DATE"), ("%m/%d/%Y",), "record date"
        )
        pay_date = _day(
            _cell(row, columns, "PAYABLE DATE"), ("%m/%d/%Y",), "pay date"
        )
        _decimal(_cell(row, columns, "DIVIDEND ($)"), "dividend", positive=True)
        if _text(_cell(row, columns, "FREQUENCY")) != "Quarterly":
            raise QualificationError("SPY distribution frequency mismatch")
        all_dates.append(ex_date)
        if start <= ex_date <= end:
            if not ex_date <= record_date <= pay_date:
                raise QualificationError("target distribution dates are not chronological")
            for field in ("SHORT TERM CAPITAL GAIN ($)", "LONG TERM CAPITAL GAIN ($)"):
                if _decimal(_text(_cell(row, columns, field)) or "0", field) != 0:
                    raise QualificationError("target period contains an unexpected capital gain")
            target_dates.append(ex_date)
    counts = Counter(value.year for value in target_dates)
    expected = Counter({year: 4 for year in range(start.year, end.year + 1)})
    if len(target_dates) != 40 or len(set(target_dates)) != 40 or counts != expected:
        raise QualificationError("target distribution coverage is not exactly four per year")
    return {
        "classification": "QUALIFIED_RETROSPECTIVE_DISTRIBUTIONS_PIT_UNQUALIFIED",
        "all_spy_event_count": len(all_dates),
        "all_spy_ex_date_min": min(all_dates).isoformat(),
        "all_spy_ex_date_max": max(all_dates).isoformat(),
        "target_event_count": 40,
        "target_year_counts": {str(year): counts[year] for year in sorted(counts)},
        "target_dates_complete": True,
        "target_amounts_complete": True,
        "historical_announced_at_complete": False,
        "revision_lineage_complete": False,
    }


def qualify_nav(raw: bytes, start: date, end: date, expected: dict[str, int]) -> dict[str, Any]:
    rows = xlsx_rows(raw)
    if len(rows) < 5 or _text(rows[1][1]) != "SPY":
        raise QualificationError("NAV workbook identity mismatch")
    header_index, columns = _header(
        rows, ("Date", "NAV", "Shares Outstanding", "Total Net Assets")
    )
    observed: dict[date, Decimal] = {}
    for row in rows[header_index + 1 :]:
        value = _text(_cell(row, columns, "Date"))
        if not value:
            continue
        try:
            day = _day(value, ("%d-%b-%Y",), "NAV date")
        except QualificationError:
            if all(not _text(_cell(row, columns, field)) for field in columns if field != "Date"):
                continue
            raise
        if day in observed:
            raise QualificationError("NAV dates must be unique")
        observed[day] = _decimal(_cell(row, columns, "NAV"), "NAV", positive=True)
        if start <= day <= end:
            _decimal(_cell(row, columns, "Shares Outstanding"), "shares", positive=True)
            _decimal(_cell(row, columns, "Total Net Assets"), "net assets", positive=True)
    target = [day for day in observed if start <= day <= end]
    counts = Counter(day.year for day in target)
    if counts != Counter({int(year): count for year, count in expected.items()}):
        raise QualificationError("target NAV session counts changed")
    return {
        "classification": "QUALIFIED_RETROSPECTIVE_NAV_NOT_TRADABLE_OHLC",
        "row_count": len(observed),
        "date_min": min(observed).isoformat(),
        "date_max": max(observed).isoformat(),
        "target_row_count": len(target),
        "target_year_counts": {str(year): counts[year] for year in sorted(counts)},
        "contains_raw_open_close": False,
    }


def qualify_premium_discount(raw: bytes, start: date, end: date) -> dict[str, Any]:
    rows = xlsx_rows(raw)
    if len(rows) < 5 or _text(rows[1][1]) != "SPY":
        raise QualificationError("premium/discount workbook identity mismatch")
    header_index, columns = _header(rows, ("Date", "Premium/Discount"))
    observed: dict[date, Decimal] = {}
    for row in rows[header_index + 1 :]:
        value = _text(_cell(row, columns, "Date"))
        if not value:
            continue
        try:
            day = _day(value, ("%d-%b-%Y",), "premium/discount date")
        except QualificationError:
            if not _text(_cell(row, columns, "Premium/Discount")):
                continue
            raise
        if day in observed:
            raise QualificationError("premium/discount dates must be unique")
        observed[day] = _decimal(_cell(row, columns, "Premium/Discount"), "premium/discount")
    target = [day for day in observed if start <= day <= end]
    return {
        "classification": "PARTIAL_RECENT_PREMIUM_DISCOUNT_NOT_RAW_OHLC",
        "row_count": len(observed),
        "date_min": min(observed).isoformat(),
        "date_max": max(observed).isoformat(),
        "target_row_count": len(target),
        "target_start_covered": min(observed) <= start,
        "target_end_covered": max(observed) >= end,
        "contains_raw_open_close": False,
    }


def qualify_taq_sample(raw: bytes) -> dict[str, Any]:
    try:
        rows = [line.split("|") for line in raw.decode("ascii").splitlines() if line]
    except UnicodeDecodeError as exc:
        raise QualificationError("TAQ sample must be ASCII") from exc
    if any(len(row) != 14 for row in rows):
        raise QualificationError("TAQ sample must contain exactly 14 fields per row")
    matches = [row for row in rows if row[0] == "SPY"]
    if len(matches) != 1 or matches[0][1] != "78462F103" or matches[0][8] != "04/01/2026":
        raise QualificationError("TAQ sample SPY identity mismatch")
    return {
        "classification": "QUALIFIED_OFFICIAL_OHLC_SCHEMA_SAMPLE_ONLY",
        "row_count": len(rows),
        "field_count": 14,
        "spy_row_count": 1,
        "sample_date": "2026-04-01",
        "target_period_covered": False,
        "historical_access": "BLOCKED_ENTITLEMENT",
    }


def qualify_source_dir(source_dir: Path, definition_path: Path = DEFINITION) -> dict[str, Any]:
    definition_raw = definition_path.read_bytes()
    definition = strict_load_bytes(definition_raw)
    manifest_raw = (source_dir / "retrieval_manifest.json").read_bytes()
    manifest = strict_load_bytes(manifest_raw)
    expected_manifest = {
        "authentication_used": False,
        "cookies_used": False,
        "redirects_followed": 0,
        "retries": 0,
    }
    if any(manifest.get(field) != value for field, value in expected_manifest.items()):
        raise QualificationError("retrieval manifest crossed the fixed network boundary")
    manifest_items = manifest.get("resources")
    if type(manifest_items) is not list or any(type(item) is not dict for item in manifest_items):
        raise QualificationError("retrieval manifest resources are invalid")
    resources = {str(item.get("source_id")): item for item in manifest_items}
    if len(resources) != len(manifest_items):
        raise QualificationError("retrieval manifest source IDs must be unique")
    expected_ids = {spec["source_id"] for spec in definition["sources"]}
    if set(resources) != expected_ids:
        raise QualificationError("retrieval manifest source scope changed")
    captured: dict[str, bytes] = {}
    for spec in definition["sources"]:
        source_id = spec["source_id"]
        record = resources.get(source_id)
        if record is None or record["url"] != spec["url"] or record["filename"] != spec["filename"]:
            raise QualificationError(f"retrieval identity mismatch: {source_id}")
        raw = (source_dir / spec["filename"]).read_bytes()
        if (
            len(raw) != record["bytes"]
            or len(raw) > spec["max_bytes"]
            or sha256_bytes(raw) != record["content_sha256"]
        ):
            raise QualificationError(f"retrieved bytes changed: {source_id}")
        available_at = _timestamp(record["observed_available_at"], "observed_available_at")
        retrieved_at = _timestamp(record["retrieved_at"], "retrieved_at")
        if available_at > retrieved_at or record.get("redirects_followed") != 0:
            raise QualificationError(f"retrieval chronology changed: {source_id}")
        captured[source_id] = raw
    start = date.fromisoformat(definition["target_period"]["start"])
    end = date.fromisoformat(definition["target_period"]["end"])
    calendars = []
    for year in range(2022, 2026):
        raw = captured[f"nyse_calendar_{year}"]
        if not raw.startswith(b"%PDF-"):
            raise QualificationError(f"NYSE {year} calendar is not a PDF")
        calendars.append({"year": year, "content_sha256": sha256_bytes(raw), "bytes": len(raw)})
    return {
        "boundaries": {
            "central_database_write_performed": False,
            "raw_values_published_in_report": False,
            "strategy_candidate_available": False,
            "strategy_execution_performed": False,
        },
        "component_results": {
            "calendar": {
                "classification": "PARTIAL_OFFICIAL_2022_2025_ONLY",
                "available_years": [2022, 2023, 2024, 2025],
                "missing_target_years": [2016, 2017, 2018, 2019, 2020, 2021],
                "documents": calendars,
            },
            "distributions": qualify_distributions(
                captured["ssga_historical_distributions"], start, end
            ),
            "nav": qualify_nav(
                captured["ssga_nav_history"],
                start,
                end,
                definition["acceptance"]["target_nav_session_counts"],
            ),
            "premium_discount": qualify_premium_discount(
                captured["ssga_premium_discount_history"], start, end
            ),
            "product_identity": qualify_product(captured["ssga_product_snapshot"]),
            "raw_ohlc": qualify_taq_sample(captured["nyse_arca_closing_prices_sample"]),
            "split_and_lifecycle": {
                "classification": "BLOCKED_COMPLETE_OFFICIAL_IDENTITY",
                "complete_2016_2025": False,
            },
        },
        "conclusion": "PARTIALLY_QUALIFIED_BLOCKED_ANONYMOUS_OFFICIAL_RAW_OHLC",
        "definition_sha256": sha256_bytes(definition_raw),
        "gate_counts": None,
        "p4_r_strict_eligible": False,
        "phase": "P3",
        "qualification_id": definition["qualification_id"],
        "residual_blockers": [
            "2016_2025_official_raw_open_close_requires_paid_entitlement",
            "complete_split_and_historical_lifecycle_identity_unavailable_anonymously",
            "official_session_calendar_missing_2016_2021",
            "historical_distribution_announced_at_and_revision_lineage_unqualified",
        ],
        "source_snapshot": {
            "manifest_sha256": sha256_bytes(manifest_raw),
            "resource_count": len(captured),
            "snapshot_id": manifest["snapshot_id"],
            "resources": {
                source_id: {
                    "bytes": resources[source_id]["bytes"],
                    "content_sha256": resources[source_id]["content_sha256"],
                    "observed_available_at": resources[source_id]["observed_available_at"],
                    "server_last_modified": resources[source_id].get("server_last_modified"),
                    "url": resources[source_id]["url"],
                }
                for source_id in sorted(captured)
            },
        },
        "strategy_candidate_available": False,
        "target_period": definition["target_period"],
        "version": 1,
    }


def write_result(result: dict[str, Any], output: Path) -> tuple[str, Path]:
    sidecar = output.with_suffix(output.suffix + ".sha256")
    if output.exists() or sidecar.exists():
        raise QualificationError("result or sidecar already exists")
    output.parent.mkdir(parents=True, exist_ok=True)
    raw = canonical_json(result)
    digest = sha256_bytes(raw)
    with tempfile.NamedTemporaryFile(dir=output.parent, delete=False) as handle:
        temporary = Path(handle.name)
        handle.write(raw)
    try:
        os.replace(temporary, output)
    finally:
        temporary.unlink(missing_ok=True)
    sidecar.write_text(f"{digest}  {output.name}\n", encoding="ascii")
    return digest, sidecar


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-dir", type=Path)
    parser.add_argument("--result", type=Path)
    parser.add_argument("--definition", type=Path, default=DEFINITION)
    args = parser.parse_args(argv)
    if args.source_dir is None and args.result is None:
        print(
            json.dumps(
                {
                    "files_written": False,
                    "network_executed": False,
                    "phase": "P3",
                    "source_count": len(strict_load(args.definition)["sources"]),
                    "status": "DRY_RUN_PLAN",
                },
                sort_keys=True,
            )
        )
        return 0
    if args.source_dir is None or args.result is None:
        raise QualificationError("--source-dir and --result must be supplied together")
    result = qualify_source_dir(args.source_dir, args.definition)
    digest, sidecar = write_result(result, args.result)
    print(
        json.dumps(
            {
                "conclusion": result["conclusion"],
                "p4_r_strict_eligible": False,
                "result_sha256": digest,
                "sidecar": str(sidecar),
                "strategy_candidate_available": False,
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
