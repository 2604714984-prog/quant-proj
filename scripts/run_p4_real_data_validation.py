#!/usr/bin/env python3
"""Validate one free-source SPY replay without claiming PIT strategy evidence."""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
from io import BytesIO
import json
import math
import os
from pathlib import Path
import stat
import subprocess
import tempfile
from typing import Any

import duckdb
from openpyxl import load_workbook

from quant_system.backtest import Portfolio, TransactionCostModel
from quant_system.data.reader import sha256_file

ROOT = Path(__file__).resolve().parents[1]
DEFINITION = ROOT / "research/definitions/p4_spy_retrospective_real_data_v1.json"
RESULT = ROOT / "research/reports/p4_spy_retrospective_real_data_v1_result.json"
RESEARCH_ID = "P4_R_SPY_FREE_SOURCE_RETROSPECTIVE_STATIC_ALLOCATION_V1"
DEFINITION_SHA256 = "a24faabaac04ede4f67b02427f8bd3f7b725b9c6f8e0dee41a167c2b8b2f8fea"
START, END = date(2016, 1, 4), date(2025, 12, 31)
GATES = (
    "definition_contract_exact",
    "runtime_and_input_hashes_exact",
    "central_database_read_only_identity",
    "sina_target_rows_complete",
    "tiingo_target_rows_complete",
    "source_dates_and_official_nav_align",
    "consumed_prices_cross_source_within_tolerance",
    "official_distributions_complete",
    "distribution_cross_source_within_tolerance",
    "reconstructed_total_return_within_tolerance",
    "next_session_entry_and_costs_exact",
    "distribution_cash_and_nav_accounting_exact",
    "deterministic_replay",
    "retrospective_boundary_preserved",
)
SCOPE = (
    "research/definitions/p4_spy_retrospective_real_data_v1.json",
    "scripts/run_p4_real_data_validation.py",
    "tests/test_p4_real_data_validation.py",
)


class ValidationError(ValueError):
    pass


def _pairs(values: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in values:
        if key in result:
            raise ValidationError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _nonfinite(value: str) -> None:
    raise ValidationError(f"nonfinite JSON constant: {value}")


def strict_load(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(
            path.read_text(encoding="utf-8"),
            object_pairs_hook=_pairs,
            parse_constant=_nonfinite,
        )
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValidationError(f"invalid JSON: {path}") from exc
    if type(value) is not dict:
        raise ValidationError("top-level JSON must be an object")
    return value


def canonical_json(value: Any) -> bytes:
    return (
        json.dumps(value, allow_nan=False, ensure_ascii=True, indent=2, sort_keys=True)
        + "\n"
    ).encode()


def _compact_json(value: Any) -> bytes:
    return json.dumps(
        value, allow_nan=False, ensure_ascii=True, separators=(",", ":"), sort_keys=True
    ).encode()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _decimal(value: Any, field: str, *, positive: bool = False) -> Decimal:
    if isinstance(value, bool):
        raise ValidationError(f"{field} must be finite")
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValidationError(f"{field} must be finite") from exc
    if not parsed.is_finite() or (positive and parsed <= 0):
        raise ValidationError(f"{field} must be finite and valid")
    return parsed


def _money(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.000000001")), "f")


def _ratio(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.000000000001")), "f")


def _contract(definition: dict[str, Any]) -> None:
    if set(definition) != {
        "boundaries", "date", "evidence_class", "inputs", "phase", "research_id",
        "result_path", "runtime_contract", "schema_version", "strategy_fixture",
        "strict_gates", "tolerances",
    }:
        raise ValidationError("definition fields changed")
    if definition["research_id"] != RESEARCH_ID:
        raise ValidationError("research identity changed")
    if (
        definition["schema_version"] != "p4-real-data-definition-v1"
        or definition["date"] != "2026-07-15"
        or definition["phase"] != "P4_R_REAL_DATA_SYSTEM_VALIDATION"
        or definition["evidence_class"]
        != "RETROSPECTIVE_NON_PIT_MULTI_SOURCE_SYSTEM_EVIDENCE"
    ):
        raise ValidationError("definition classification changed")
    if tuple(definition["strict_gates"]) != GATES:
        raise ValidationError("strict gates changed")
    if definition["result_path"] != str(RESULT.relative_to(ROOT)):
        raise ValidationError("result path changed")
    expected_boundaries = {
        "central_database_write_allowed": False,
        "network_allowed": False,
        "strategy_candidate_available": False,
        "strategy_evidence_eligible": False,
        "strict_pit_eligible": False,
    }
    if definition["boundaries"] != expected_boundaries:
        raise ValidationError("retrospective boundaries changed")
    if definition["runtime_contract"] != {
        "clean_committed_worktree_required_for_execute": True,
        "default_mode": "validate_only",
        "deterministic_result_bytes": True,
        "output_overwrite_allowed": False,
        "read_only_database": True,
    }:
        raise ValidationError("runtime contract changed")
    if definition["strategy_fixture"] != {
        "commission_bps": "5",
        "distribution_policy": "cash_on_pay_date_no_reinvestment",
        "entry_session": "2016-01-05",
        "final_mark_session": "2025-12-31",
        "initial_cash": "100000",
        "position_policy": "whole_shares_residual_cash",
        "signal_session": "2016-01-04",
        "slippage_bps": "5",
        "symbol": "SPY",
        "target_weight": "1",
    }:
        raise ValidationError("strategy fixture changed")
    inputs = definition["inputs"]
    if type(inputs) is not dict or set(inputs) != {
        "central_database_sha256",
        "p3_evidence",
        "sina",
        "state_street_distributions",
        "state_street_nav",
        "tiingo",
    }:
        raise ValidationError("input fields changed")
    p3 = inputs["p3_evidence"]
    if type(p3) is not dict or p3 != {
        "manifest_relative_path": (
            "raw/official/spy/ssga/ssga_spy_20260715T052647Z/"
            "retrieval_manifest.json"
        ),
        "manifest_sha256": p3.get("manifest_sha256"),
        "result_relative_path": (
            "research/reports/p3_spy_official_source_qualification_v1_result.json"
        ),
        "result_sha256": p3.get("result_sha256"),
    }:
        raise ValidationError("P3 evidence contract changed")
    sina = inputs["sina"]
    if type(sina) is not dict or set(sina) != {
        "classification", "ordered_row_hash_sha256", "row_count", "snapshot_id",
        "source_document_sha256", "table",
    } or (
        sina["classification"] != "RETROSPECTIVE_RAW_OHLC_AVAILABLE_AT_UNKNOWN"
        or sina["row_count"] != 2514
        or sina["table"] != "us_equity_research.sina_daily_bars"
    ):
        raise ValidationError("Sina contract changed")
    tiingo = inputs["tiingo"]
    if type(tiingo) is not dict or set(tiingo) != {
        "classification", "ordered_row_hash_sha256", "row_count", "snapshot_id",
        "source_document_sha256", "table",
    } or (
        tiingo["classification"] != "RETROSPECTIVE_ONLY_UNQUALIFIED_PIT_DIAGNOSTIC"
        or tiingo["row_count"] != 2514
        or tiingo["table"]
        != "us_equity_research.us_daily_total_return_research"
    ):
        raise ValidationError("Tiingo contract changed")
    distributions = inputs["state_street_distributions"]
    if type(distributions) is not dict or distributions != {
        "classification": "QUALIFIED_RETROSPECTIVE_DISTRIBUTIONS_PIT_UNQUALIFIED",
        "content_sha256": distributions.get("content_sha256"),
        "event_count": 40,
        "event_identity_sha256": distributions.get("event_identity_sha256"),
        "relative_path": (
            "raw/official/spy/ssga/ssga_spy_20260715T052647Z/"
            "spdr-etf-historical-distributions.xlsx"
        ),
    }:
        raise ValidationError("distribution contract changed")
    nav = inputs["state_street_nav"]
    if type(nav) is not dict or nav != {
        "classification": "OFFICIAL_SESSION_DATE_CROSS_CHECK_ONLY",
        "content_sha256": nav.get("content_sha256"),
        "relative_path": (
            "raw/official/spy/ssga/ssga_spy_20260715T052647Z/navhist-us-en-spy.xlsx"
        ),
        "target_row_count": 2514,
    }:
        raise ValidationError("NAV contract changed")
    if definition["tolerances"] != {
        "distribution_amount_absolute_usd": "0.005",
        "internal_accounting_absolute_usd": "0.00000001",
        "raw_ohlc_absolute_usd": "0.05",
        "total_return_cumulative_relative": "0.005",
        "total_return_daily_factor_absolute": "0.005",
    }:
        raise ValidationError("tolerances changed")
    hashes = (
        inputs["central_database_sha256"], p3["manifest_sha256"],
        p3["result_sha256"], sina["ordered_row_hash_sha256"],
        sina["source_document_sha256"], distributions["content_sha256"],
        distributions["event_identity_sha256"], nav["content_sha256"],
        tiingo["ordered_row_hash_sha256"], tiingo["source_document_sha256"],
    )
    if any(
        type(value) is not str
        or len(value) != 64
        or any(char not in "0123456789abcdef" for char in value)
        for value in hashes
    ):
        raise ValidationError("invalid frozen SHA-256")
    if any(_decimal(value, "tolerance") < 0 for value in definition["tolerances"].values()):
        raise ValidationError("negative tolerance")


def _runtime_identity(require_clean: bool) -> dict[str, Any]:
    def git(*arguments: str) -> str:
        return subprocess.run(
            ("git", "-C", str(ROOT), *arguments),
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()

    dirty = git("status", "--porcelain=v1")
    if require_clean and dirty:
        raise ValidationError("--execute requires a clean committed worktree")
    if require_clean:
        for path in SCOPE:
            git("ls-files", "--error-unmatch", path)
    return {
        "clean_committed": not dirty,
        "commit": git("rev-parse", "HEAD^{commit}"),
        "tree": git("rev-parse", "HEAD^{tree}"),
    }


def _pinned_bytes(root: Path, relative: str, expected_sha: str) -> bytes:
    root = root.expanduser().resolve()
    candidate = root / relative
    resolved = candidate.resolve()
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise ValidationError("external input escapes data root") from exc
    flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(candidate, flags)
    except OSError as exc:
        raise ValidationError(f"cannot open pinned input: {relative}") from exc
    try:
        if not stat.S_ISREG(os.fstat(descriptor).st_mode):
            raise ValidationError(f"pinned input is not a regular file: {relative}")
        chunks: list[bytes] = []
        while True:
            chunk = os.read(descriptor, 1024 * 1024)
            if not chunk:
                break
            chunks.append(chunk)
    finally:
        os.close(descriptor)
    raw = b"".join(chunks)
    if sha256_bytes(raw) != expected_sha:
        raise ValidationError(f"external input hash changed: {relative}")
    return raw


def _xlsx_rows(raw: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise ValidationError("invalid XLSX") from exc
    try:
        if len(workbook.sheetnames) != 1:
            raise ValidationError("XLSX must contain one worksheet")
        return [list(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _header(rows: list[list[Any]], required: tuple[str, ...]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows):
        columns = {_text(value): index for index, value in enumerate(row) if _text(value)}
        if all(name in columns for name in required):
            return row_index, columns
    raise ValidationError(f"XLSX header missing: {required}")


def _cell(row: list[Any], columns: dict[str, int], name: str) -> Any:
    index = columns[name]
    return row[index] if index < len(row) else None


def _excel_day(value: Any, field: str, pattern: str) -> date:
    if type(value) is datetime:
        return value.date()
    if type(value) is date:
        return value
    try:
        return datetime.strptime(str(value).strip(), pattern).date()
    except ValueError as exc:
        raise ValidationError(f"invalid {field}") from exc


def _distributions(raw: bytes) -> list[dict[str, str]]:
    rows = _xlsx_rows(raw)
    required = (
        "TICKER", "CUSIP", "EX-DATE", "RECORD DATE", "PAYABLE DATE", "DIVIDEND ($)",
        "SHORT TERM CAPITAL GAIN ($)", "LONG TERM CAPITAL GAIN ($)", "FREQUENCY",
    )
    header, columns = _header(rows, required)
    result: list[dict[str, str]] = []
    for row in rows[header + 1 :]:
        if _text(_cell(row, columns, "TICKER")) != "SPY":
            continue
        if _text(_cell(row, columns, "CUSIP")) != "78462F103":
            raise ValidationError("SPY CUSIP changed")
        ex_date = _excel_day(_cell(row, columns, "EX-DATE"), "ex-date", "%m/%d/%Y")
        if not START <= ex_date <= END:
            continue
        record_date = _excel_day(
            _cell(row, columns, "RECORD DATE"), "record date", "%m/%d/%Y"
        )
        pay_date = _excel_day(
            _cell(row, columns, "PAYABLE DATE"), "pay date", "%m/%d/%Y"
        )
        amount = _decimal(_cell(row, columns, "DIVIDEND ($)"), "dividend", positive=True)
        if not ex_date <= record_date <= pay_date:
            raise ValidationError("distribution date order changed")
        if _text(_cell(row, columns, "FREQUENCY")) != "Quarterly":
            raise ValidationError("distribution frequency changed")
        for field in ("SHORT TERM CAPITAL GAIN ($)", "LONG TERM CAPITAL GAIN ($)"):
            if _decimal(_text(_cell(row, columns, field)) or "0", field) != 0:
                raise ValidationError("unexpected capital-gain distribution")
        result.append(
            {
                "ex_date": ex_date.isoformat(),
                "record_date": record_date.isoformat(),
                "pay_date": pay_date.isoformat(),
                "amount": str(amount),
            }
        )
    return sorted(result, key=lambda item: item["ex_date"])


def _nav_dates(raw: bytes) -> list[date]:
    rows = _xlsx_rows(raw)
    if len(rows) < 3 or _text(rows[1][1]) != "SPY":
        raise ValidationError("NAV workbook identity changed")
    header, columns = _header(rows, ("Date", "NAV", "Shares Outstanding", "Total Net Assets"))
    dates: list[date] = []
    for row in rows[header + 1 :]:
        raw = _text(_cell(row, columns, "Date"))
        if not raw:
            continue
        try:
            day = _excel_day(raw, "NAV date", "%d-%b-%Y")
        except ValidationError:
            continue
        if START <= day <= END:
            dates.append(day)
    if len(dates) != len(set(dates)):
        raise ValidationError("duplicate NAV date")
    return sorted(dates)


def _query_rows(
    connection: duckdb.DuckDBPyConnection,
    statement: str,
    parameters: tuple[Any, ...],
) -> list[dict[str, Any]]:
    cursor = connection.execute(statement, parameters)
    columns = [item[0] for item in cursor.description]
    rows = cursor.fetchmany(3001)
    if len(rows) > 3000:
        raise ValidationError("source query was truncated")
    return [dict(zip(columns, row, strict=True)) for row in rows]


def _load_sources(database: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    connection = duckdb.connect(str(database), read_only=True)
    try:
        connection.execute("BEGIN TRANSACTION")
        sina = _query_rows(
            connection,
            """SELECT snapshot_id, symbol, trade_date, open, high, low, close, volume,
                      source_document_sha256, available_at, row_sha256, quality_status
                 FROM us_equity_research.sina_daily_bars
                WHERE symbol=? AND trade_date BETWEEN ? AND ? ORDER BY trade_date""",
            ("SPY", START, END),
        )
        tiingo = _query_rows(
            connection,
            """SELECT snapshot_id, symbol, trade_date, open, high, low, close, volume,
                      adj_open, adj_high, adj_low, adj_close, div_cash, split_factor,
                      gross_return_factor, adjusted_reference_factor,
                      source_document_sha256, available_at, availability_basis,
                      row_sha256, quality_status
                 FROM us_equity_research.us_daily_total_return_research
                WHERE symbol=? AND trade_date BETWEEN ? AND ? ORDER BY trade_date""",
            ("SPY", START, END),
        )
        connection.execute("ROLLBACK")
        return sina, tiingo
    finally:
        connection.close()


def _row_identity(rows: list[dict[str, Any]]) -> str:
    return sha256_bytes("".join(str(row["row_sha256"]) for row in rows).encode())


def _valid_rows(rows: list[dict[str, Any]], expected: dict[str, Any], tiingo: bool) -> bool:
    if len(rows) != expected["row_count"] or len({row["trade_date"] for row in rows}) != len(rows):
        return False
    if not rows or rows[0]["trade_date"] != START or rows[-1]["trade_date"] != END:
        return False
    if {row["snapshot_id"] for row in rows} != {expected["snapshot_id"]}:
        return False
    if {row["source_document_sha256"] for row in rows} != {expected["source_document_sha256"]}:
        return False
    if _row_identity(rows) != expected["ordered_row_hash_sha256"]:
        return False
    for row in rows:
        if row["available_at"] is not None:
            return False
        for field in ("open", "high", "low", "close"):
            value = row[field]
            if not isinstance(value, (float, int)) or not math.isfinite(value) or value <= 0:
                return False
        if not isinstance(row["volume"], int) or row["volume"] < 0:
            return False
        if tiingo:
            for field in ("adj_open", "adj_high", "adj_low", "adj_close", "split_factor"):
                value = row[field]
                if not isinstance(value, (float, int)) or not math.isfinite(value) or value <= 0:
                    return False
            if row["availability_basis"] != "UNKNOWN":
                return False
            if row["quality_status"] != "PASS_RETROSPECTIVE_ONLY_UNQUALIFIED_PIT":
                return False
        elif row["quality_status"] != "PASS":
            return False
    return True


def _product(values: list[Decimal]) -> Decimal:
    result = Decimal("1")
    for value in values:
        result *= value
    return result


def _evaluate(
    definition: dict[str, Any],
    database: Path,
    data_root: Path,
    runtime: dict[str, Any],
) -> dict[str, Any]:
    inputs = definition["inputs"]
    dist = inputs["state_street_distributions"]
    nav = inputs["state_street_nav"]
    p3 = inputs["p3_evidence"]
    distribution_bytes = _pinned_bytes(
        data_root, dist["relative_path"], dist["content_sha256"]
    )
    nav_bytes = _pinned_bytes(data_root, nav["relative_path"], nav["content_sha256"])
    _pinned_bytes(data_root, p3["manifest_relative_path"], p3["manifest_sha256"])
    _pinned_bytes(ROOT, p3["result_relative_path"], p3["result_sha256"])
    database_hash_before = sha256_file(database)
    if database_hash_before != inputs["central_database_sha256"]:
        raise ValidationError("central database SHA-256 changed")
    sina, tiingo = _load_sources(database)
    database_hash_after = sha256_file(database)
    if database_hash_after != database_hash_before:
        raise ValidationError("central database changed during read-only snapshot")
    if not sina or len(sina) != len(tiingo):
        raise ValidationError("market source row counts changed")
    distributions = _distributions(distribution_bytes)
    nav_dates = _nav_dates(nav_bytes)
    dates = [row["trade_date"] for row in sina]
    tiingo_dates = [row["trade_date"] for row in tiingo]
    event_identity = sha256_bytes(_compact_json(distributions))
    years = Counter(date.fromisoformat(item["ex_date"]).year for item in distributions)

    ohlc_max = max(
        abs(_decimal(left[field], field) - _decimal(right[field], field))
        for left, right in zip(sina, tiingo, strict=True)
        for field in ("open", "high", "low", "close")
    )
    official = {
        date.fromisoformat(item["ex_date"]): _decimal(item["amount"], "amount")
        for item in distributions
    }
    provider = {
        row["trade_date"]: _decimal(row["div_cash"] or 0, "div_cash")
        for row in tiingo
        if _decimal(row["div_cash"] or 0, "div_cash") > 0
    }
    shared = sorted(set(official) & set(provider))
    amount_max = max(
        (abs(official[day] - provider[day]) for day in shared),
        default=Decimal("Infinity"),
    )

    sina_factors, tiingo_factors = [], []
    daily_max = Decimal("0")
    for index in range(1, len(sina)):
        day = sina[index]["trade_date"]
        left = (
            _decimal(sina[index]["close"], "close") + official.get(day, Decimal("0"))
        ) / _decimal(sina[index - 1]["close"], "previous close", positive=True)
        right = _decimal(tiingo[index]["gross_return_factor"], "gross factor", positive=True)
        sina_factors.append(left)
        tiingo_factors.append(right)
        daily_max = max(daily_max, abs(left - right))
    sina_total, tiingo_total = _product(sina_factors), _product(tiingo_factors)
    cumulative_error = abs(sina_total - tiingo_total) / abs(tiingo_total)

    fixture = definition["strategy_fixture"]
    initial = _decimal(fixture["initial_cash"], "initial cash", positive=True)
    commission_rate = _decimal(fixture["commission_bps"], "commission") / Decimal("10000")
    slippage_rate = _decimal(fixture["slippage_bps"], "slippage") / Decimal("10000")
    entry_day = date.fromisoformat(fixture["entry_session"])
    entry_index = dates.index(entry_day)
    entry_open = _decimal(sina[entry_index]["open"], "entry open", positive=True)
    fill = entry_open * (Decimal("1") + slippage_rate)
    shares = int(initial / (fill * (Decimal("1") + commission_rate)))
    portfolio = Portfolio.us(
        float(initial), costs=TransactionCostModel(commission_rate=float(commission_rate))
    )
    portfolio.start_session(entry_day)
    trade = portfolio.buy("SPY", shares, float(fill), entry_day)
    events = {date.fromisoformat(item["ex_date"]): item for item in distributions}
    entitlement_total, applied = Decimal("0"), 0
    for day in dates[entry_index:]:
        portfolio.start_session(day)
        event = events.get(day)
        if event:
            entitlement_total += _decimal(
                portfolio.apply_cash_distribution(
                    "SPY",
                    event_id=f"SSGA-SPY-{event['ex_date']}",
                    amount_per_share=float(_decimal(event["amount"], "amount", positive=True)),
                    ex_date=day,
                    pay_date=date.fromisoformat(event["pay_date"]),
                ),
                "entitlement",
            )
            applied += 1
    final_close = _decimal(sina[-1]["close"], "final close", positive=True)
    final_nav = _decimal(portfolio.nav({"SPY": float(final_close)}), "final NAV")
    commission = _decimal(trade.costs.commission, "commission")
    residual = initial - Decimal(shares) * fill - commission
    direct_nav = residual + entitlement_total + Decimal(shares) * final_close
    accounting_error = abs(final_nav - direct_nav)

    tolerance = {key: _decimal(value, key) for key, value in definition["tolerances"].items()}
    ohlc_conflicts = [
        (left["trade_date"], field, difference)
        for left, right in zip(sina, tiingo, strict=True)
        for field in ("open", "high", "low", "close")
        if (
            difference := abs(
                _decimal(left[field], field) - _decimal(right[field], field)
            )
        ) > tolerance["raw_ohlc_absolute_usd"]
    ]
    consumed_price_max = max(
        abs(
            _decimal(sina[entry_index]["open"], "Sina entry open")
            - _decimal(tiingo[entry_index]["open"], "Tiingo entry open")
        ),
        abs(
            _decimal(sina[-1]["close"], "Sina final close")
            - _decimal(tiingo[-1]["close"], "Tiingo final close")
        ),
    )
    expected_pending = sum(
        (
            Decimal(shares) * _decimal(item["amount"], "pending amount", positive=True)
            for item in distributions
            if date.fromisoformat(item["ex_date"]) >= entry_day
            and date.fromisoformat(item["pay_date"]) > END
        ),
        Decimal("0"),
    )
    pending_cash = _decimal(portfolio.pending_cash_total, "pending cash")
    settled_cash = _decimal(portfolio.settled_cash, "settled cash")
    pending_error = abs(pending_cash - expected_pending)
    settled_error = abs(
        settled_cash - (residual + entitlement_total - expected_pending)
    )
    runtime_hashes_exact = (
        runtime.get("clean_committed") is True
        and all(
            type(runtime.get(field)) is str
            and len(runtime[field]) == 40
            and all(character in "0123456789abcdef" for character in runtime[field])
            for field in ("commit", "tree")
        )
    )
    gates = {
        "definition_contract_exact": True,
        "runtime_and_input_hashes_exact": runtime_hashes_exact,
        "central_database_read_only_identity": (
            database_hash_before == database_hash_after == inputs["central_database_sha256"]
        ),
        "sina_target_rows_complete": _valid_rows(sina, inputs["sina"], False),
        "tiingo_target_rows_complete": _valid_rows(tiingo, inputs["tiingo"], True),
        "source_dates_and_official_nav_align": dates == tiingo_dates == nav_dates,
        "consumed_prices_cross_source_within_tolerance": (
            consumed_price_max <= tolerance["raw_ohlc_absolute_usd"]
        ),
        "official_distributions_complete": (
            len(distributions) == dist["event_count"]
            and event_identity == dist["event_identity_sha256"]
            and years == Counter({year: 4 for year in range(2016, 2026)})
        ),
        "distribution_cross_source_within_tolerance": (
            set(official) == set(provider)
            and amount_max <= tolerance["distribution_amount_absolute_usd"]
        ),
        "reconstructed_total_return_within_tolerance": (
            cumulative_error <= tolerance["total_return_cumulative_relative"]
            and daily_max <= tolerance["total_return_daily_factor_absolute"]
        ),
        "next_session_entry_and_costs_exact": (
            dates[0] == START
            and dates[1] == entry_day
            and trade.shares == shares
            and shares > 0
        ),
        "distribution_cash_and_nav_accounting_exact": (
            accounting_error <= tolerance["internal_accounting_absolute_usd"]
            and applied == 40
            and pending_error <= tolerance["internal_accounting_absolute_usd"]
            and settled_error <= tolerance["internal_accounting_absolute_usd"]
        ),
        "deterministic_replay": True,
        "retrospective_boundary_preserved": (
            all(row["available_at"] is None for row in sina + tiingo)
            and definition["boundaries"]["strict_pit_eligible"] is False
            and definition["boundaries"]["strategy_evidence_eligible"] is False
        ),
    }
    return {
        "gate_results": gates,
        "portfolio": {
            "commission": _money(commission),
            "distribution_cash_entitled": _money(entitlement_total),
            "distribution_cash_pending_receivable": _money(pending_cash),
            "distribution_cash_settled": _money(
                entitlement_total - expected_pending
            ),
            "entry_fill_price": _money(fill),
            "final_close": _money(final_close),
            "final_nav": _money(final_nav),
            "initial_cash": _money(initial),
            "residual_cash_after_entry": _money(residual),
            "shares": shares,
            "total_return": _ratio(final_nav / initial - Decimal("1")),
        },
        "reconciliation": {
            "distribution_amount_max_absolute_usd": _money(amount_max),
            "distribution_dates_matched": len(shared),
            "consumed_price_max_absolute_usd": _money(consumed_price_max),
            "raw_ohlc_conflict_cell_count": len(ohlc_conflicts),
            "raw_ohlc_conflict_session_count": len(
                {item[0] for item in ohlc_conflicts}
            ),
            "raw_ohlc_max_absolute_usd": _money(ohlc_max),
            "sina_ssga_total_return_factor": _ratio(sina_total),
            "tiingo_total_return_factor": _ratio(tiingo_total),
            "total_return_cumulative_relative_error": _ratio(cumulative_error),
            "total_return_daily_factor_max_absolute_error": _ratio(daily_max),
        },
        "source_evidence": {
            "database_sha256": database_hash_after,
            "distribution_event_identity_sha256": event_identity,
            "nav_date_count": len(nav_dates),
            "sina_ordered_row_hash_sha256": _row_identity(sina),
            "sina_row_count": len(sina),
            "tiingo_ordered_row_hash_sha256": _row_identity(tiingo),
            "tiingo_row_count": len(tiingo),
        },
    }


def build_report(
    definition_path: Path,
    database: Path,
    data_root: Path,
    *,
    runtime: dict[str, Any],
) -> dict[str, Any]:
    definition = strict_load(definition_path)
    _contract(definition)
    observed = _evaluate(definition, database, data_root, runtime)
    observed["gate_results"]["deterministic_replay"] = observed == _evaluate(
        definition, database, data_root, runtime
    )
    failed = [name for name in GATES if observed["gate_results"].get(name) is not True]
    return {
        "boundary_result": "PASS_RETROSPECTIVE_NON_PIT_ONLY" if not failed else "FAIL_CLOSED",
        "evidence_class": definition["evidence_class"],
        "gate_results": observed["gate_results"],
        "identities": {
            "definition_sha256": sha256_path(definition_path),
            "git_commit": runtime["commit"],
            "git_tree": runtime["tree"],
            "runner_sha256": sha256_path(Path(__file__).resolve()),
            "tests_sha256": sha256_path(ROOT / SCOPE[2]),
        },
        "phase": definition["phase"],
        "portfolio": observed["portfolio"],
        "reconciliation": observed["reconciliation"],
        "research_id": RESEARCH_ID,
        "schema_version": "p4-real-data-result-v1",
        "source_evidence": observed["source_evidence"],
        "strategy_candidate_available": False,
        "strategy_evidence_eligible": False,
        "strategy_gate_counts": None,
        "strict_pit_eligible": False,
        "system_gate_counts": {"passed": len(GATES) - len(failed), "total": len(GATES)},
        "system_validation_status": (
            "P4_R_RETROSPECTIVE_SYSTEM_VALIDATION_PASS"
            if not failed
            else "P4_R_RETROSPECTIVE_SYSTEM_VALIDATION_FAIL"
        ),
    }


def _publish(report: dict[str, Any], output: Path) -> tuple[str, Path]:
    sidecar = output.with_suffix(output.suffix + ".sha256")
    if output.exists() or sidecar.exists():
        raise ValidationError("result or sidecar already exists")
    output.parent.mkdir(parents=True, exist_ok=True)
    raw = canonical_json(report)
    digest = sha256_bytes(raw)
    items = ((sidecar, f"{digest}  {output.name}\n".encode()), (output, raw))
    staged: list[tuple[Path, str]] = []
    published: list[Path] = []
    try:
        for path, content in items:
            descriptor, temporary = tempfile.mkstemp(
                prefix=f".{path.name}.", dir=path.parent
            )
            staged.append((path, temporary))
            with os.fdopen(descriptor, "wb") as handle:
                handle.write(content)
                handle.flush()
                os.fsync(handle.fileno())
        for path, temporary in staged:
            os.replace(temporary, path)
            published.append(path)
            directory = os.open(
                path.parent, os.O_RDONLY | getattr(os, "O_DIRECTORY", 0)
            )
            try:
                os.fsync(directory)
            finally:
                os.close(directory)
    except BaseException:
        for path in reversed(published):
            path.unlink(missing_ok=True)
        raise
    finally:
        for _, temporary in staged:
            if os.path.exists(temporary):
                os.unlink(temporary)
    return digest, sidecar


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--definition", type=Path, default=DEFINITION)
    parser.add_argument("--data-root", type=Path)
    parser.add_argument("--database", type=Path)
    parser.add_argument("--output", type=Path, default=RESULT)
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args(argv)
    definition_path = args.definition.expanduser().resolve()
    if (
        definition_path != DEFINITION.resolve()
        or sha256_path(definition_path) != DEFINITION_SHA256
    ):
        raise ValidationError("canonical definition path or SHA-256 changed")
    if args.output.expanduser().resolve() != RESULT.resolve():
        raise ValidationError("canonical result path changed")
    definition = strict_load(definition_path)
    _contract(definition)
    if not args.execute:
        print(json.dumps({
            "database_read": False,
            "files_written": False,
            "mode": "VALIDATE_ONLY",
            "network_executed": False,
            "research_id": RESEARCH_ID,
        }, sort_keys=True))
        return 0
    data_root_value = args.data_root or os.environ.get("QUANT_DATA_ROOT")
    if data_root_value is None:
        raise ValidationError("--data-root or QUANT_DATA_ROOT is required")
    data_root = Path(data_root_value).expanduser().resolve()
    database = (args.database or data_root / "quant_research.duckdb").expanduser().resolve()
    try:
        database.relative_to(data_root)
    except ValueError as exc:
        raise ValidationError("database must remain under data root") from exc
    runtime = _runtime_identity(require_clean=True)
    report = build_report(definition_path, database, data_root, runtime=runtime)
    digest, sidecar = _publish(report, args.output)
    print(json.dumps({
        "result": str(args.output),
        "result_sha256": digest,
        "sidecar": str(sidecar),
        "status": report["system_validation_status"],
    }, sort_keys=True))
    return 0 if report["system_gate_counts"]["passed"] == len(GATES) else 2


if __name__ == "__main__":
    raise SystemExit(main())
